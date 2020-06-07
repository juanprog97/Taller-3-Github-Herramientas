from openpyxl import load_workbook
import xlsxwriter
import sys
from PyQt5 import QtCore, QtGui, QtWidgets
from PyQt5.QtWidgets import QApplication, QWidget, QInputDialog, QLineEdit, QFileDialog
from PyQt5.QtGui import QIcon
import threading

class asignRolPerfectly():
    def __init__(self,rol,source,search):
        self.sourceDoc = load_workbook(filename=source)
        self.searchDoc = load_workbook(filename=search)
        self.rol = rol
        self.sheetSource = self.sourceDoc.active
        self.sheetSearch = self.searchDoc.active


        self.RolandCharge = []
        self.NotFounded = [] 

        self.loadDataSource()
        self.loadDataSearch()

    def loadDataSource(self):
        self.inputData = []
        for value in self.sheetSource.iter_rows(min_row=2,min_col=2,max_col=3,values_only=True):
            self.inputData.append((str(value[0]).strip(),value[1]))
        print(len(self.inputData))
    
    def loadDataSearch(self):
        self.dataUserOfficial = {}
        for value in self.sheetSearch.iter_rows(min_row=2,min_col=1,max_col=4,values_only=True):
            self.dataUserOfficial[str(value[0]).strip()] = [value[2],value[3]]
        print(len(self.dataUserOfficial))
    



class Ui_MainPrincipal(QWidget):
    def setupUi(self, MainPrincipal):
        MainPrincipal.setObjectName("MainPrincipal")
        MainPrincipal.resize(469, 231)
        sizePolicy = QtWidgets.QSizePolicy(QtWidgets.QSizePolicy.Minimum, QtWidgets.QSizePolicy.Preferred)
        sizePolicy.setHorizontalStretch(0)
        sizePolicy.setVerticalStretch(0)
        sizePolicy.setHeightForWidth(MainPrincipal.sizePolicy().hasHeightForWidth())
        MainPrincipal.setSizePolicy(sizePolicy)
        self.centralwidget = QtWidgets.QWidget(MainPrincipal)
        self.centralwidget.setObjectName("centralwidget")
        self.GenerarDoc = QtWidgets.QPushButton(self.centralwidget)
        self.GenerarDoc.setGeometry(QtCore.QRect(190, 180, 81, 31))
        self.GenerarDoc.setObjectName("GenerarDoc")
        self.GenerarDoc.clicked.connect(self.generarDoc)
        self.GenerarDoc.setEnabled(False)
        self.verticalLayoutWidget = QtWidgets.QWidget(self.centralwidget)
        self.verticalLayoutWidget.setGeometry(QtCore.QRect(40, 30, 371, 81))
        self.verticalLayoutWidget.setObjectName("verticalLayoutWidget")
        self.verticalLayout = QtWidgets.QVBoxLayout(self.verticalLayoutWidget)
        self.verticalLayout.setContentsMargins(0, 0, 0, 0)
        self.verticalLayout.setSpacing(3)
        self.verticalLayout.setObjectName("verticalLayout")
        self.label_2 = QtWidgets.QLabel(self.verticalLayoutWidget)
        font = QtGui.QFont()
        font.setPointSize(9)
        font.setBold(True)
        font.setWeight(75)
        self.label_2.setFont(font)
        self.label_2.setObjectName("label_2")
        self.verticalLayout.addWidget(self.label_2)
        self.horizontalLayout = QtWidgets.QHBoxLayout()
        self.horizontalLayout.setObjectName("horizontalLayout")
        self.textEdit = QtWidgets.QTextEdit(self.verticalLayoutWidget)
        self.textEdit.setVerticalScrollBarPolicy(QtCore.Qt.ScrollBarAlwaysOff)
        self.textEdit.setUndoRedoEnabled(False)
        self.textEdit.setReadOnly(True)
        self.textEdit.setObjectName("textEdit")
        self.horizontalLayout.addWidget(self.textEdit)
        self.buscarDocListado = QtWidgets.QPushButton(self.verticalLayoutWidget)
        self.buscarDocListado.setObjectName("buscarDocListado")
        self.buscarDocListado.clicked.connect(self.clickedButtonListadoAgregados)#call the function for open de file picker
        self.horizontalLayout.addWidget(self.buscarDocListado)
        self.verticalLayout.addLayout(self.horizontalLayout)
        self.label_3 = QtWidgets.QLabel(self.verticalLayoutWidget)
        font = QtGui.QFont()
        font.setPointSize(9)
        font.setBold(True)
        font.setWeight(75)
        self.label_3.setFont(font)
        self.label_3.setObjectName("label_3")
        self.verticalLayout.addWidget(self.label_3)
        self.horizontalLayout_3 = QtWidgets.QHBoxLayout()
        self.horizontalLayout_3.setObjectName("horizontalLayout_3")
        self.textEdit_3 = QtWidgets.QTextEdit(self.verticalLayoutWidget)
        self.textEdit_3.setVerticalScrollBarPolicy(QtCore.Qt.ScrollBarAlwaysOff)
        self.textEdit_3.setUndoRedoEnabled(False)
        self.textEdit_3.setReadOnly(True)
        self.textEdit_3.setObjectName("textEdit_3")
        self.horizontalLayout_3.addWidget(self.textEdit_3)
        self.buscarDocExac = QtWidgets.QPushButton(self.verticalLayoutWidget)
        self.buscarDocExac.setObjectName("buscarDocExac")
        self.buscarDocExac.clicked.connect(self.clickedButtonMatrizCargos) #call the function for open de file picker
        self.buscarDocExac.setEnabled(False)
        self.horizontalLayout_3.addWidget(self.buscarDocExac)
        self.verticalLayout.addLayout(self.horizontalLayout_3)
        self.nombreRol = QtWidgets.QPlainTextEdit(self.centralwidget)
        self.nombreRol.setGeometry(QtCore.QRect(40, 130, 281, 21))
        self.nombreRol.setVerticalScrollBarPolicy(QtCore.Qt.ScrollBarAlwaysOff)
        self.nombreRol.setObjectName("nombreRol")
        self.nombreRol.setEnabled(False)
        self.label = QtWidgets.QLabel(self.centralwidget)
        self.label.setGeometry(QtCore.QRect(330, 130, 111, 21))
        font = QtGui.QFont()
        font.setPointSize(10)
        font.setBold(True)
        font.setWeight(75)
        self.label.setFont(font)
        self.label.setObjectName("label")
        MainPrincipal.setCentralWidget(self.centralwidget)

        self.retranslateUi(MainPrincipal)
        QtCore.QMetaObject.connectSlotsByName(MainPrincipal)


    def openFileNameDialog1(self):
        options = QFileDialog.Options()
        options |= QFileDialog.DontUseNativeDialog
        fileName, _ = QFileDialog.getOpenFileName(self,"Directorio", "","All Files (*);;Excel Files (*.xlsx)", options=options)
        if fileName:
            self.source = fileName
            self.textEdit.setText(str(fileName))
            self.buscarDocExac.setEnabled(True)


    def openFileNameDialog2(self):
        options = QFileDialog.Options()
        options |= QFileDialog.DontUseNativeDialog
        fileName, _ = QFileDialog.getOpenFileName(self,"Directorio", "","All Files (*);;Excel Files (*.xlsx)", options=options)
        if fileName:
            self.search = fileName
            self.textEdit_3.setText(str(fileName))
            print(self.search)
            self.nombreRol.setEnabled(True)
            self.GenerarDoc.setEnabled(True)


    def  clickedButtonMatrizCargos(self):   #Function Create QTwigdetTrre
        self.openFileNameDialog2()

    def clickedButtonListadoAgregados(self): #Function Create QTwigdetTrre
        self.openFileNameDialog1()
    
    def generarDoc(self):
        
        self.analizedDoc = asignRolPerfectly(self.nombreRol.toPlainText(),self.source,self.search)
        #self.analizedDoc.rolVsCharge()
        
    def retranslateUi(self, MainPrincipal):
        _translate = QtCore.QCoreApplication.translate
        MainPrincipal.setWindowTitle(_translate("MainPrincipal", "Asignador De Roles"))
        self.GenerarDoc.setText(_translate("MainPrincipal", "Generar"))
        self.label_2.setText(_translate("MainPrincipal", "Listado de usuarios por asignar"))
        self.buscarDocListado.setText(_translate("MainPrincipal", "Browser"))
        self.label_3.setText(_translate("MainPrincipal", "Matriz de Cargos IDM"))
        self.buscarDocExac.setText(_translate("MainPrincipal", "Browser"))
        self.label.setText(_translate("MainPrincipal", "Nombre del Rol"))

if __name__ == "__main__":
    app = QtWidgets.QApplication(sys.argv)
    MainPrincipal = QtWidgets.QMainWindow()
    ui = Ui_MainPrincipal()
    ui.setupUi(MainPrincipal)

    MainPrincipal.show()
    
    sys.exit(app.exec_())